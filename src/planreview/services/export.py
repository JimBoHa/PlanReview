from __future__ import annotations

import json
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from sqlmodel import select

from planreview.config import get_settings
from planreview.database import session_scope
from planreview.models import Discrepancy, UploadedDocument


def _project_export_dir(project_id: str) -> Path:
    settings = get_settings()
    export_dir = settings.projects_dir / project_id / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _build_excel(project_id: str, rows: list[Discrepancy]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Discrepancies"
    sheet.append(
        [
            "Row",
            "Source",
            "Page",
            "Citation",
            "Description",
            "Thumbnail",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row.row_number,
                row.document_kind,
                row.page_label,
                row.citation,
                row.description,
                row.thumbnail_path,
            ]
        )
        if row.thumbnail_path:
            image_path = (
                get_settings().projects_dir
                / project_id
                / "artifacts"
                / Path(row.thumbnail_path).name
            )
            if image_path.exists():
                image = Image(str(image_path))
                image.width = 140
                image.height = 100
                anchor = f"F{sheet.max_row}"
                sheet.add_image(image, anchor)
                sheet.row_dimensions[sheet.max_row].height = 80
    path = _project_export_dir(project_id) / "PlanReview-comments.xlsx"
    workbook.save(path)
    return path


def _apply_markup(input_path: str, output_path: Path, rows: list[Discrepancy]) -> None:
    rows_by_page: dict[int, list[Discrepancy]] = {}
    for row in rows:
        rows_by_page.setdefault(row.page_number, []).append(row)
    with fitz.open(input_path) as document:
        for page_number, findings in rows_by_page.items():
            page = document[page_number - 1]
            for finding in findings:
                rect = fitz.Rect(json.loads(finding.markup_rect))
                annotation = page.add_rect_annot(rect)
                annotation.set_colors(stroke=(1, 0, 0))
                annotation.update()
                page.insert_text(
                    point=(rect.x1 + 8, max(rect.y0, 20)),
                    text=f"#{finding.row_number}",
                    fontsize=11,
                    color=(1, 0, 0),
                )
        document.save(output_path)


def export_project_outputs(project_id: str) -> dict[str, str]:
    with session_scope() as session:
        rows = session.exec(
            select(Discrepancy)
            .where(Discrepancy.project_id == project_id)
            .order_by(Discrepancy.row_number)
        ).all()
        documents = session.exec(
            select(UploadedDocument).where(UploadedDocument.project_id == project_id)
        ).all()

    export_dir = _project_export_dir(project_id)
    outputs: dict[str, str] = {"excel": str(_build_excel(project_id, list(rows)))}
    for document in documents:
        relevant = [row for row in rows if row.document_id == document.id]
        output_name = f"PlanReview-{document.kind.value}-marked.pdf"
        path = export_dir / output_name
        _apply_markup(document.stored_path, path, relevant)
        outputs[document.kind.value] = str(path)
    return outputs
